import pandas as pd
import time
from bs4 import BeautifulSoup
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

import smtplib
from email.message import EmailMessage

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException


# Функция возвращает правильное склонение слова "строка"
def get_word(n):
    d = n % 10
    if d == 0 or d >= 5:
        return 'строк'
    if d == 1:
        return 'строка'
    return 'строки'


# Функция задаёт поля фильтра за последний месяц
def set_data_of_prev_month():
    global driver

    # Получение даты - последний день предыдущего месяца
    prev_month = date.today().replace(day=1) - timedelta(days=1)

    # Определение полей фильтра - день, месяц
    d1day_select = Select(driver.find_element(By.ID, 'd1day'))
    d2day_select = Select(driver.find_element(By.ID, 'd2day'))
    d1month_select = Select(driver.find_element(By.ID, 'd1month'))
    d2month_select = Select(driver.find_element(By.ID, 'd2month'))

    # Установка полей фильтра - день, месяц
    d1month_select.select_by_value(str(prev_month.month))
    d2month_select.select_by_value(str(prev_month.month))
    d1day_select.select_by_value('1')
    d2day_select.select_by_value(str(prev_month.day))

    # Установка предыдущего года, если предыдущий месяц - декабрь
    if prev_month.month == 12:
        d1year_select = Select(driver.find_element(By.ID, 'd1year'))
        d2year_select = Select(driver.find_element(By.ID, 'd2year'))
        d1year_select.select_by_value(str(prev_month.year))
        d2year_select.select_by_value(str(prev_month.year))

    # Нажать на кнопку Показать
    submit_button = driver.find_element(By.NAME, 'bSubmit')
    submit_button.click()


# Функция излечения данных таблицы из вёрстки. Возвращает датафрейм
def get_dataframe(table_bs):
    data = []

    for tr in table_bs.find_all('tr'):
        row = []
        for td in tr.find_all('td'):
            row.append(td.get_text())
        if row:
            data.append(row)

    df = pd.DataFrame(data, columns=['Дата', '-', '-', 'Курс', 'Время'])
    del df['-']
    return df


# Путь до файла Excel
path_excel_file = '../data/excel.xlsx'

# Инициализация драйвера браузера
driver = webdriver.Firefox(executable_path='H:/Python/Parsing/code/geckodriver.exe')

# Переход на сайт
driver.get('https://www.moex.com/')

# Открыть меню (окно браузера должно быть поверх всех остальных)
menu_button = driver.find_element(By.CSS_SELECTOR,
                                  'span.header-menu__item button.header-menu__link')
menu_button.click()

# Открыть раздел Срочный рынок
futures_market_a = menu_button.find_elements(By.XPATH, "//a[contains(text(),'Срочный рынок')]")
futures_market_a = [fm for fm in futures_market_a if fm.text != ''][0]
futures_market_a.click()

# Согласиться с условиями Пользовательского соглашения
try:
    agree_a = driver.find_element(By.LINK_TEXT, 'Согласен')
    if agree_a:
        agree_a.click()
except NoSuchElementException:
    pass

# Пауза. Если не поставить, то будет ошибка при переходе в Индикативные курсы:
# Элемент не доступен для клика в точке ..., потому что другой элемент ... закрывает его
time.sleep(2)

# Перейти в раздел Индикативные курсы
currency_rate_a = driver.find_element(By.LINK_TEXT, 'Индикативные курсы')
currency_rate_a.click()

# Запрос курса USD за прошлый месяц
currency_select = Select(driver.find_element(By.ID, 'ctl00_PageContent_CurrencySelect'))
currency_select.select_by_visible_text('USD/RUB - Доллар США к российскому рублю')
set_data_of_prev_month()

# Извлечение таблицы с историей курса USD
tablels_usd = driver.find_element(By.CLASS_NAME, 'tablels')
table_usd_bs = BeautifulSoup(tablels_usd.get_attribute('innerHTML'))

# Запрос курса JPY за прошлый месяц
currency_select = Select(driver.find_element(By.ID, 'ctl00_PageContent_CurrencySelect'))
currency_select.select_by_visible_text('JPY/RUB - Японская йена к российскому рублю')
set_data_of_prev_month()

# Извлечение таблицы с историей курса JPY
tablels_jpy = driver.find_element(By.CLASS_NAME, 'tablels')
table_jpy_bs = BeautifulSoup(tablels_jpy.get_attribute('innerHTML'))

# Закрытие браузера и формирование таблиц (датафреймов)
driver.close()
df_usd = get_dataframe(table_usd_bs)
df_jpy = get_dataframe(table_jpy_bs)

# Преборазование строк в числа
for df in [df_usd, df_jpy]:
    df['Курс'] = df['Курс'].str.replace(',', '.').astype(float)

# Создание книги Excel и получение текущего листа
wb = Workbook()
ws = wb.active

# Запись данных курса USD на лист с заголовками
for row in dataframe_to_rows(df_usd, index=False, header=True):
    ws.append(row)

# Запись заголовков на листт для данных JPY
for i, col in enumerate(['D', 'E', 'F']):
    ws[col + '1'] = df_jpy.columns[i]

# Запись данных курса JPY на лист
for i, row in df_jpy.iterrows():
    for j, col in enumerate(['D', 'E', 'F']):
        ws[col + str(i + 2)] = row[j]

# Создание столбца Результат и вычисление значений
ws['G1'] = 'Результат'
for i in range(2, ws.max_row + 1):
    si = str(i)
    ws['G' + si] = float(ws['B' + si].value) / float(ws['E' + si].value)

# Задать автоширину для столбцов
for i in range(1, ws.max_column + 1):
    ws.column_dimensions[get_column_letter(i)].bestFit = True
    ws.column_dimensions[get_column_letter(i)].auto_size = True

# Сохранение файла Excel
wb.save(path_excel_file)

# Определение отправителя, получателя, тему, текст письма
msg = EmailMessage()
sender = 'kruglik.a.s@mail.ru'
recipients = ['MikAlBelov@Greenatom.ru', sender]
msg['From'] = sender
msg['To'] = recipients
msg['Subject'] = 'Круглик Андрей Сергеевич. Тестовое задание на Python'
message = str(ws.max_row) + ' ' + get_word(ws.max_row) + \
        ' (без заголовков: ' + str(ws.max_row - 1) + ' ' + get_word(ws.max_row - 1) + ')'
msg.set_content(message, 'plain')

# Прикрепление файла к письму
with open(path_excel_file, 'rb') as f:
    file_data = f.read()
    msg.add_attachment(file_data, maintype="application", 
                       subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                       filename='excel.xlsx')

# Отправка письма
with smtplib.SMTP_SSL('smtp.mail.ru', 465) as mailserver:
    mailserver.login(sender, 'password_for_app')
    mailserver.sendmail(sender, recipients, msg.as_string())
